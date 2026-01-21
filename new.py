import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, Border, Side, Alignment

#  1. Load CSV into pandas and Preprocess ---
df = pd.read_csv("Auto Sales data.csv")
df['ORDERDATE'] = pd.to_datetime(df['ORDERDATE'], format='%d/%m/%Y')

#  2. Generate Summary Stats (Step 6) ---
sales_stats = df['SALES'].agg(['count', 'mean', 'std', 'min', 'max', 'sum']).to_frame(name='SALES')
# FIX: Explicitly cast to object/string type to avoid FutureWarning
sales_stats['SALES'] = sales_stats['SALES'].astype(object)
# ... rest of the formatting ...
# Format the statistics for the report
for stat in ['mean', 'std', 'min', 'max', 'sum']:
    sales_stats.loc[stat, 'SALES'] = f"${sales_stats.loc[stat, 'SALES']:,.2f}"
sales_stats.loc['count', 'SALES'] = f"{int(sales_stats.loc['count', 'SALES']):,}"

#  3. Create Pivot Tables (Step 2) ---

# Pivot 1: Total Sales by PRODUCTLINE
pivot_product_sales = df.pivot_table(
    index='PRODUCTLINE',
    values='SALES',
    aggfunc='sum'
).sort_values(by='SALES', ascending=False)
pivot_product_sales['Total Sales ($)'] = pivot_product_sales['SALES'].apply(lambda x: f"${x:,.2f}")
pivot_product_sales = pivot_product_sales.drop(columns=['SALES'])


# Pivot 2: Average Sales by COUNTRY and DEALSIZE
pivot_country_dealsize = df.pivot_table(
    index='COUNTRY',
    columns='DEALSIZE',
    values='SALES',
    aggfunc='mean'
).round(2).fillna(0)
# Format the values for the report
for col in pivot_country_dealsize.columns:
    pivot_country_dealsize[col] = pivot_country_dealsize[col].apply(lambda x: f"${x:,.2f}")

# --- 4. Generate Chart (Step 3) ---
chart_data = df.pivot_table(
    index='PRODUCTLINE',
    values='SALES',
    aggfunc='sum'
).sort_values(by='SALES', ascending=False)

plt.figure(figsize=(10, 6))
chart_data['SALES'].plot(kind='bar', color='darkblue')
plt.title('Total Sales by Product Line', fontsize=16)
plt.ylabel('Total Sales ($)', fontsize=12)
plt.xlabel('Product Line', fontsize=12)
plt.xticks(rotation=45, ha='right')
plt.grid(axis='y', linestyle='--', alpha=0.7)
plt.tight_layout()
chart_filename = 'sales_by_productline_chart.png'
plt.savefig(chart_filename)
plt.close()


# --- 5. Export and Style Excel (Step 4) ---

excel_file = 'Auto_Sales_Report.xlsx'
writer = pd.ExcelWriter(excel_file, engine='openpyxl')

# Write DataFrames to different sheets
sales_stats.to_excel(writer, sheet_name='Summary', startrow=0, header=False)
pivot_product_sales.to_excel(writer, sheet_name='Summary', startrow=10)
pivot_country_dealsize.to_excel(writer, sheet_name='Country_Analysis', startrow=0)

writer.close() # Close the Pandas writer to save the dataframes

# Load workbook for advanced styling and adding the chart
wb = load_workbook(excel_file)
ws_summary = wb['Summary']
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# 5.1. Add titles and insert chart on 'Summary' sheet
ws_summary['A1'] = 'Overall Sales Statistics'
ws_summary['A1'].font = Font(bold=True, size=14)
ws_summary['A1'].alignment = Alignment(horizontal='left')

ws_summary['A10'] = 'Sales by Product Line'
ws_summary['A10'].font = Font(bold=True, size=14)

img = OpenpyxlImage(chart_filename)
img.width = 600
img.height = 360
ws_summary.add_image(img, 'D12') # Anchor chart in cell D12

# 5.2. Apply basic border styling to tables
def apply_border_style(ws, start_row, end_row, start_col, end_col):
    for r_idx in range(start_row, end_row + 1):
        for c_idx in range(start_col, end_col + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = thin_border
            if r_idx == start_row:
                 cell.font = Font(bold=True)
            # Right align numerical data columns
            if c_idx > start_col:
                cell.alignment = Alignment(horizontal='right')

# Style Sales Stats Table (A2:B7)
apply_border_style(ws_summary, 2, 7, 1, 2)
# Style Sales by Product Line Pivot Table (A11:B18)
apply_border_style(ws_summary, 11, 18, 1, 2)
# Style Country Analysis Sheet
ws_country = wb['Country_Analysis']
apply_border_style(ws_country, 1, len(pivot_country_dealsize) + 1, 1, len(pivot_country_dealsize.columns) + 1)

# Set column widths for better readability
ws_summary.column_dimensions['A'].width = 20
ws_summary.column_dimensions['B'].width = 18
ws_country.column_dimensions['A'].width = 18

wb.save(excel_file)

print(f"Excel Report saved as {excel_file}")
